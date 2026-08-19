"""按人事模板填写月度上下班与支援工时。"""

from __future__ import annotations

from calendar import monthrange
from copy import copy
from datetime import date
from decimal import Decimal
from io import BytesIO
from collections import defaultdict

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import Session, joinedload

from app.models import Employee, WorkEntry
from app.services.entries import entry_hours_decimal
from app.services.hr_export_clock import clock_in_out_for_entry, excel_hour_value
from app.services.hr_excel_layout import (
    DAY1_COL,
    DAY_HEADER_ROW,
    LABEL_COL,
    MONTH_DATE_COL,
    MONTH_DATE_ROW,
    NAME_COL,
    PERSON_START_ROW,
    POSITION_COL,
    ROWS_PER_PERSON,
    SEQ_COL,
    STORE_COL,
    STORE_ROW,
    SUPPORT_VALUE_COL,
    TEMPLATE_PATH,
    TEMPLATE_PERSON_SLOTS,
    WEEKDAY_ROW,
)
from app.services.settings import get_store_name

WEEKDAY_LABELS = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]


def export_filename(store_name: str, month: int) -> str:
    return f"{store_name}店{month}月份.xlsx"


def list_export_employees(db: Session, year: int, month: int) -> list[Employee]:
    month_start = date(year, month, 1)
    month_end = date(year, month, monthrange(year, month)[1])
    emp_ids = (
        select(WorkEntry.employee_id)
        .where(
            WorkEntry.work_date >= month_start,
            WorkEntry.work_date <= month_end,
        )
        .distinct()
    )
    return list(
        db.scalars(
            select(Employee)
            .where(Employee.id.in_(emp_ids))
            .order_by(Employee.sort_order, Employee.created_at)
        ).all()
    )


def _copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)
        dst.alignment = copy(src.alignment)


def _set_cell(ws: Worksheet, row: int, col: int, value) -> None:
    """Set value only on writable cells (skip MergedCell bottoms)."""
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def _ensure_person_rows(ws: Worksheet, person_index: int) -> int:
    """Return 0-indexed on-duty row for person_index; append styled rows if needed."""
    on_row = PERSON_START_ROW + person_index * ROWS_PER_PERSON
    if person_index < TEMPLATE_PERSON_SLOTS:
        return on_row

    prev_on = PERSON_START_ROW + (person_index - 1) * ROWS_PER_PERSON
    max_col = ws.max_column
    for offset in (0, 1):
        src_r = prev_on + offset + 1
        dst_r = on_row + offset + 1
        if dst_r > ws.max_row:
            ws.append([])
        for col in range(1, max_col + 1):
            src = ws.cell(src_r, col)
            dst = ws.cell(dst_r, col)
            if isinstance(dst, MergedCell):
                continue
            _copy_cell_style(src, dst)
            if col - 1 == LABEL_COL:
                dst.value = src.value
            else:
                dst.value = None
        if src_r in ws.row_dimensions:
            ws.row_dimensions[dst_r].height = ws.row_dimensions[src_r].height
    # Merge identity cols like template (A/B/C across on+off)
    on_r = on_row + 1
    off_r = on_row + 2
    for col0 in (SEQ_COL, NAME_COL, POSITION_COL):
        col = col0 + 1
        merge_range = f"{ws.cell(on_r, col).coordinate}:{ws.cell(off_r, col).coordinate}"
        if merge_range not in [str(m) for m in ws.merged_cells.ranges]:
            try:
                ws.merge_cells(start_row=on_r, start_column=col, end_row=off_r, end_column=col)
            except ValueError:
                pass
    return on_row


def _clear_person_slot(
    ws: Worksheet, on_row0: int, days_in_month: int, person_index: int
) -> None:
    on_r = on_row0 + 1
    off_r = on_row0 + 2
    _set_cell(ws, on_r, SEQ_COL + 1, None)
    _set_cell(ws, on_r, NAME_COL + 1, None)
    _set_cell(ws, on_r, POSITION_COL + 1, None)
    for r in (on_r, off_r):
        for d in range(31):
            _set_cell(ws, r, DAY1_COL + 1 + d, None)
        # AN3 是模板里的“支援”标题（第一位 on 行）。
        # 当 person_index==0 且 r==on_r 时，跳过清空以避免覆盖标题。
        if not (person_index == 0 and r == on_r):
            _set_cell(ws, r, SUPPORT_VALUE_COL + 1, None)


def _write_day_headers(ws: Worksheet, year: int, month: int) -> int:
    days = monthrange(year, month)[1]
    for d in range(1, 32):
        col = DAY1_COL + d  # 1-based openpyxl col for day d
        if d <= days:
            ws.cell(DAY_HEADER_ROW + 1, col).value = d
            weekday = date(year, month, d).weekday()
            ws.cell(WEEKDAY_ROW + 1, col).value = WEEKDAY_LABELS[weekday]
        else:
            ws.cell(DAY_HEADER_ROW + 1, col).value = None
            ws.cell(WEEKDAY_ROW + 1, col).value = None
    return days


def build_export_workbook(db: Session, year: int, month: int) -> bytes:
    month_start = date(year, month, 1)
    month_end = date(year, month, monthrange(year, month)[1])

    entries = (
        db.scalars(
            select(WorkEntry)
            .options(joinedload(WorkEntry.employee))
            .where(
                WorkEntry.work_date >= month_start,
                WorkEntry.work_date <= month_end,
            )
        )
        .unique()
        .all()
    )

    by_emp: dict = defaultdict(list)
    for e in entries:
        by_emp[e.employee_id].append(e)

    people = list_export_employees(db, year, month)

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    store_name = get_store_name(db)
    ws.cell(STORE_ROW + 1, STORE_COL + 1).value = store_name
    # 只改值，保留已有日期格式
    ws.cell(MONTH_DATE_ROW + 1, MONTH_DATE_COL + 1).value = month_start

    days = _write_day_headers(ws, year, month)

    for i in range(TEMPLATE_PERSON_SLOTS):
        on_row0 = PERSON_START_ROW + i * ROWS_PER_PERSON
        if i >= len(people):
            _clear_person_slot(ws, on_row0, days, i)
            continue
        _fill_person(ws, i, people[i], by_emp.get(people[i].id, []), days)

    # Extra people beyond template slots
    for i in range(TEMPLATE_PERSON_SLOTS, len(people)):
        _ensure_person_rows(ws, i)
        _fill_person(ws, i, people[i], by_emp.get(people[i].id, []), days)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fill_person(
    ws: Worksheet,
    index: int,
    emp: Employee,
    entries: list[WorkEntry],
    days_in_month: int,
) -> None:
    on_row0 = PERSON_START_ROW + index * ROWS_PER_PERSON
    on_r = on_row0 + 1
    off_r = on_row0 + 2

    display_name = emp.export_name or emp.name
    _set_cell(ws, on_r, SEQ_COL + 1, index + 1)
    _set_cell(ws, on_r, NAME_COL + 1, display_name)
    _set_cell(ws, on_r, POSITION_COL + 1, emp.position)

    by_day = {e.work_date.day: e for e in entries}

    for d in range(1, 32):
        col = DAY1_COL + d
        if d > days_in_month or d not in by_day:
            _set_cell(ws, on_r, col, None)
            _set_cell(ws, off_r, col, None)
            continue
        entry = by_day[d]
        inn, out = clock_in_out_for_entry(entry)
        _set_cell(ws, on_r, col, excel_hour_value(inn) if inn is not None else None)
        _set_cell(ws, off_r, col, excel_hour_value(out) if out is not None else None)

    support_sum = sum(
        (entry_hours_decimal(e) for e in entries if e.status == "support"),
        Decimal("0"),
    )
    support_col = SUPPORT_VALUE_COL + 1

    # “支援”标题占用第一位人员 on 行（AN3），数值需要写到 AN4。
    if index == 0:
        if support_sum > 0:
            _set_cell(ws, off_r, support_col, excel_hour_value(support_sum))
        else:
            # 保持 AN3 标题，保证 AN4 为空
            _set_cell(ws, off_r, support_col, None)
        # 不触碰 on_r（AN3 标题）
    else:
        if support_sum > 0:
            _set_cell(ws, on_r, support_col, excel_hour_value(support_sum))
        else:
            _set_cell(ws, on_r, support_col, None)
        _set_cell(ws, off_r, support_col, None)
