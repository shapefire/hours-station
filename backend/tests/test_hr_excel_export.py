from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from app.services.hr_excel_export import build_export_workbook, export_filename
from app.services.hr_excel_layout import (
    DAY1_COL,
    LABEL_COL,
    NAME_COL,
    PERSON_START_ROW,
    POSITION_COL,
    RATE_COL,
    ROWS_PER_PERSON,
    SEQ_COL,
    SUPPORT_VALUE_COL,
    TEMPLATE_EXTENDED_SLOTS,
    TEMPLATE_PERSON_SLOTS,
    TEMPLATE_PATH,
    TOTAL_HEADER_COL,
    TRIPLE_PAY_COL,
    WEEKDAY_ROW,
)


def test_export_filename():
    assert export_filename("东圃地铁站", 7) == "东圃地铁站店7月份.xlsx"


def test_february_headers_and_blank_extra_days(db):
    raw = build_export_workbook(db, 2026, 2)
    ws = load_workbook(BytesIO(raw)).active
    day1_col = DAY1_COL + 1
    assert ws.cell(1, day1_col).value == 1
    # 2026-02-01 周日
    assert ws.cell(WEEKDAY_ROW + 1, day1_col).value == "周日"
    assert ws.cell(1, day1_col + 27).value == 28
    assert ws.cell(1, day1_col + 28).value in (None, "")
    assert ws.cell(WEEKDAY_ROW + 1, day1_col + 28).value in (None, "")


def test_only_registered_people_in_sort_order(client, db):
    zhang = client.post("/api/employees", json={"name": "张三"}).json()
    li = client.post("/api/employees", json={"name": "李四"}).json()
    client.put("/api/employees/reorder", json={"ids": [li["id"], zhang["id"]]})
    client.patch(
        f"/api/employees/{li['id']}",
        json={"export_name": "李四全", "position": "全职"},
    )
    client.post(
        "/api/entries",
        json={
            "work_date": "2026-08-01",
            "name": "李四",
            "start_time": "07:30",
            "end_time": "16:00",
        },
    )
    # 张三当月无登记
    raw = build_export_workbook(db, 2026, 8)
    ws = load_workbook(BytesIO(raw)).active
    first_name_row = PERSON_START_ROW + 1  # openpyxl 1-based
    assert ws.cell(first_name_row, NAME_COL + 1).value == "李四全"
    assert ws.cell(first_name_row, POSITION_COL + 1).value == "全职"
    assert ws.cell(first_name_row, SEQ_COL + 1).value == 1
    # 下一人员上班行（空槽）
    second_on_row = PERSON_START_ROW + ROWS_PER_PERSON + 1
    assert ws.cell(second_on_row, NAME_COL + 1).value in (None, "")
    assert ws.cell(second_on_row, SEQ_COL + 1).value in (None, "")
    assert ws.cell(second_on_row, LABEL_COL + 1).value == "上班"
    assert ws.cell(second_on_row + 1, LABEL_COL + 1).value == "下班"


def test_duty_cells_and_support_column(client, db):
    client.post(
        "/api/entries",
        json={
            "work_date": "2026-08-01",
            "name": "苑菱",
            "start_time": "07:30",
            "end_time": "16:00",
            "ot_start_time": "22:00",
            "ot_end_time": "23:30",
        },
    )
    raw = build_export_workbook(db, 2026, 8)
    ws = load_workbook(BytesIO(raw)).active
    on_cell = ws.cell(PERSON_START_ROW + 1, DAY1_COL + 1).value
    off_cell = ws.cell(PERSON_START_ROW + 2, DAY1_COL + 1).value
    assert on_cell in (8, 8.0)
    assert off_cell in (17.5, 17.50)
    # 无支援 → 支援列空
    # 模板在 AN3 放了“支援”标题；第一位人员的支援数值应在 AN4
    on_row = PERSON_START_ROW + 1  # openpyxl 1-based
    off_row = on_row + 1
    an_col = 40  # openpyxl 列号：AN
    assert ws.cell(off_row, an_col).value in (None, "")


def test_rest_with_ot_and_pure_rest(client, db):
    client.post(
        "/api/entries",
        json={
            "work_date": "2026-08-01",
            "name": "甲",
            "status": "rest",
            "ot_start_time": "22:00",
            "ot_end_time": "23:30",
        },
    )
    client.post(
        "/api/entries",
        json={
            "work_date": "2026-08-02",
            "name": "甲",
            "status": "rest",
        },
    )
    raw = build_export_workbook(db, 2026, 8)
    ws = load_workbook(BytesIO(raw)).active
    on_row = PERSON_START_ROW + 1
    off_row = PERSON_START_ROW + 2
    # 休息+加班：上下班为加班起止小时
    assert ws.cell(on_row, DAY1_COL + 1).value in (22, 22.0)
    assert ws.cell(off_row, DAY1_COL + 1).value in (23.5, 23.50)
    # 纯休息：空
    assert ws.cell(on_row, DAY1_COL + 1 + 1).value in (None, "")
    assert ws.cell(off_row, DAY1_COL + 1 + 1).value in (None, "")


def test_support_day_and_support_column(client, db):
    client.post(
        "/api/entries",
        json={
            "work_date": "2026-08-05",
            "name": "乙",
            "status": "support",
            "start_time": "11:00",
            "end_time": "19:00",
        },
    )
    raw = build_export_workbook(db, 2026, 8)
    ws = load_workbook(BytesIO(raw)).active
    on_row = PERSON_START_ROW + 1
    off_row = PERSON_START_ROW + 2
    day5_col = DAY1_COL + 1 + 4  # 公历 5 日
    # 支援日与到岗同样填上下班
    assert ws.cell(on_row, day5_col).value in (11.5, 11.50)
    assert ws.cell(off_row, day5_col).value in (19, 19.0)
    # 支援列 = 有效工时 8-0.5=7.5
    # 第一位人员的支援数值应在 AN4
    off_row = on_row + 1
    an_col = 40  # openpyxl 列号：AN
    support_val = ws.cell(off_row, an_col).value
    assert support_val in (7.5, 7.50)


def test_store_name_and_month_anchor(db):
    raw = build_export_workbook(db, 2026, 8)
    ws = load_workbook(BytesIO(raw)).active
    from app.services.hr_excel_layout import MONTH_DATE_COL, STORE_COL, STORE_ROW

    assert ws.cell(STORE_ROW + 1, STORE_COL + 1).value == "东圃地铁站"
    month_val = ws.cell(STORE_ROW + 1, MONTH_DATE_COL + 1).value
    if hasattr(month_val, "date"):
        month_val = month_val.date()
    assert month_val == date(2026, 8, 1)


def test_on_duty_short_shift_no_deduction_in_export_cells(client, db):
    """
    主时段扣减只应在命中扣减档时发生；
    小于 min_hours 的班次不应仍然在导出里额外扣减（当前实现会写死 +0.5，故此用例会失败）。
    """
    client.post(
        "/api/entries",
        json={
            "work_date": "2026-08-01",
            "name": "短班",
            "start_time": "08:00",
            "end_time": "11:30",
        },
    )
    raw = build_export_workbook(db, 2026, 8)
    ws = load_workbook(BytesIO(raw)).active

    on_row = PERSON_START_ROW + 1
    off_row = PERSON_START_ROW + 2
    day1_col = DAY1_COL + 1

    # 预期：导出里的上班/下班差值应等于系统 effective_hours（无 OT）。
    from datetime import time
    from app.services.hours import effective_hours
    from app.services.hr_export_clock import time_to_hour_number, excel_hour_value

    start_t = time(8, 0)
    end_t = time(11, 30)
    effective_main = effective_hours(start_t, end_t)
    end_h = time_to_hour_number(end_t)
    expected_on = end_h - effective_main
    expected_off = end_h

    on_cell = ws.cell(on_row, day1_col).value
    off_cell = ws.cell(off_row, day1_col).value

    expected_on_val = excel_hour_value(expected_on)
    expected_off_val = excel_hour_value(expected_off)

    assert on_cell in (expected_on_val, float(expected_on_val))
    assert off_cell in (expected_off_val, float(expected_off_val))


def test_extra_person_beyond_template_slots_has_summary_formulas(client, db):
    """超过模板 15 人槽位时，动态追加行应补写總計/Rate 公式（不含三薪）。"""
    for i in range(TEMPLATE_PERSON_SLOTS + 1):
        client.post(
            "/api/entries",
            json={
                "work_date": "2026-08-01",
                "name": f"员工{i + 1}",
                "start_time": "08:00",
                "end_time": "16:00",
            },
        )

    raw = build_export_workbook(db, 2026, 8)
    ws = load_workbook(BytesIO(raw)).active

    extra_index = TEMPLATE_PERSON_SLOTS
    on_row = PERSON_START_ROW + extra_index * ROWS_PER_PERSON + 1
    off_row = on_row + 1
    aj_col = TOTAL_HEADER_COL + 1
    al_col = RATE_COL + 1
    am_col = TRIPLE_PAY_COL + 1

    assert ws.cell(on_row, aj_col).value == (
        f"=SUM(E{off_row}:AI{off_row})-SUM(E{on_row}:AI{on_row})"
    )
    assert ws.cell(on_row, al_col).value == f"=COUNTIF(E{off_row}:AI{off_row},22.5)"
    assert ws.cell(on_row, am_col).value in (None, "")
    assert ws.cell(on_row, NAME_COL + 1).value == f"员工{TEMPLATE_PERSON_SLOTS + 1}"
    assert f"AJ{on_row}:AJ{off_row}" in [str(m) for m in ws.merged_cells.ranges]


def test_within_template_slots_do_not_add_rate_triple_pay_formulas(client, db):
    client.post(
        "/api/entries",
        json={
            "work_date": "2026-08-01",
            "name": "张三",
            "start_time": "08:00",
            "end_time": "16:00",
        },
    )
    raw = build_export_workbook(db, 2026, 8)
    ws = load_workbook(BytesIO(raw)).active
    on_row = PERSON_START_ROW + 1
    assert ws.cell(on_row, RATE_COL + 1).value == "总工时"
    assert ws.cell(on_row, TRIPLE_PAY_COL + 1).value == "三薪"
    # 第二位人员不再保留模板三薪 COUNTIF
    second_on_row = PERSON_START_ROW + ROWS_PER_PERSON + 1
    assert ws.cell(second_on_row, TRIPLE_PAY_COL + 1).value in (None, "")


def test_extended_template_slot_without_new_row_copy_has_summary_formulas(client, db):
    """模板 16–23 号槽（rows 33–48）已有行结构，导出时补總計/Rate 公式，不统计三薪。"""
    for i in range(TEMPLATE_EXTENDED_SLOTS):
        client.post(
            "/api/entries",
            json={
                "work_date": "2026-08-01",
                "name": f"员工{i + 1}",
                "start_time": "08:00",
                "end_time": "16:00",
            },
        )

    raw = build_export_workbook(db, 2026, 8)
    ws = load_workbook(BytesIO(raw)).active

    last_index = TEMPLATE_EXTENDED_SLOTS - 1
    on_row = PERSON_START_ROW + last_index * ROWS_PER_PERSON + 1
    off_row = on_row + 1
    aj_col = TOTAL_HEADER_COL + 1

    assert ws.cell(on_row, aj_col).value == (
        f"=SUM(E{off_row}:AI{off_row})-SUM(E{on_row}:AI{on_row})"
    )
    assert ws.cell(on_row, RATE_COL + 1).value == f"=COUNTIF(E{off_row}:AI{off_row},22.5)"
    assert ws.cell(on_row, TRIPLE_PAY_COL + 1).value in (None, "")


def test_template_total_formulas_match_full_day_range():
    """
    模板“總計/AJ”公式的日列范围应从 E 列开始（第 1 日列），并一直覆盖到 AI 列。
    """
    wb = load_workbook(TEMPLATE_PATH, data_only=False)
    ws = wb.active

    aj_col = TOTAL_HEADER_COL + 1  # openpyxl 1-based

    for person_index in range(TEMPLATE_PERSON_SLOTS):
        on_row = PERSON_START_ROW + person_index * ROWS_PER_PERSON + 1
        off_row = on_row + 1

        expected = f"=SUM(E{off_row}:AI{off_row})-SUM(E{on_row}:AI{on_row})"
        assert ws.cell(on_row, aj_col).value == expected
